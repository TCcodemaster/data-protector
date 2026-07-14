#!/usr/bin/env python3
"""PreToolUse hook handler for data-protector plugin.

Intercepts Read/Bash/R-Studio tool calls that read data files.
- If columns not configured yet: block and trigger column selection flow
- If columns configured: block and suggest filtered command

Key rule: ANY Bash command referencing a data file path is blocked,
regardless of how it reads the file (cat, python script, heredoc, etc.)
"""
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_manager import load_config, is_file_configured, CONFIG_PATH

PLUGIN_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILTER_SCRIPT = os.path.join(PLUGIN_ROOT, "scripts", "filter.py")
LAST_FILE_PATH = os.path.expanduser("~/.claude/data-protector-last-file.txt")

DATA_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".parquet"}
DATA_EXT_PATTERN = re.compile(
    r"\.(csv|tsv|xlsx|xls|parquet)\b", re.IGNORECASE
)

SQL_PATTERNS = re.compile(
    r"\b(sqlite3|psql|mysql|duckdb)\b.*\bSELECT\b", re.IGNORECASE
)
COPY_CMDS = re.compile(r"\b(cp|ln|mv|copy|symlink)\b")
R_PATTERNS = re.compile(
    r"(read\.csv|read\.table|read\.delim|readr::read_csv|readr::read_tsv|"
    r"data\.table::fread|fread|read_excel|readxl::read_excel|"
    r"read\.xlsx|openxlsx::read\.xlsx)"
)

HEADER_READ_MARKER = "__dp_header_read__"


def is_already_filtered(command):
    return "filter.py" in command


def is_header_read(command):
    return HEADER_READ_MARKER in command


def has_data_file_ref(text):
    return bool(DATA_EXT_PATTERN.search(text))


def extract_data_file_path(text):
    patterns = [
        re.compile(r"""['"]([^'"]*\.(csv|tsv|xlsx|xls|json|parquet))['"]""", re.IGNORECASE),
        re.compile(r"""(\S+\.(csv|tsv|xlsx|xls|json|parquet))\b""", re.IGNORECASE),
    ]
    for p in patterns:
        m = p.search(text)
        if m:
            return m.group(1)
    return None


def resolve_file_path(tool_input, tool_name):
    if tool_name == "Read":
        return tool_input.get("file_path", "")
    elif tool_name == "Bash":
        return extract_data_file_path(tool_input.get("command", ""))
    return None


def read_headers(file_path):
    """Read column headers from data file. Returns list (flat) or dict (per-sheet for xlsx)."""
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext in (".csv", ".tsv"):
            import csv
            delim = "\t" if ext == ".tsv" else ","
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter=delim)
                return next(reader, [])
        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
            except ImportError:
                return []
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = {}
            for name in wb.sheetnames:
                ws = wb[name]
                first_row = next(ws.iter_rows(max_row=1), None)
                if first_row:
                    sheets[name] = [str(c.value) if c.value is not None else "" for c in first_row]
            wb.close()
            if len(sheets) == 1:
                return list(sheets.values())[0]
            return sheets
        elif ext == ".json":
            import json as json_mod
            with open(file_path, "r", encoding="utf-8") as f:
                data = json_mod.load(f)
            if isinstance(data, list) and data and isinstance(data[0], dict):
                return list(data[0].keys())
            elif isinstance(data, dict):
                return list(data.keys())
        return []
    except Exception:
        return []



def save_last_file(file_path):
    """Record last blocked file so /protect can target it."""
    if file_path:
        try:
            abs_path = os.path.abspath(file_path)
            os.makedirs(os.path.dirname(LAST_FILE_PATH), exist_ok=True)
            with open(LAST_FILE_PATH, "w") as f:
                f.write(abs_path)
        except Exception:
            pass


def block_with_filter(reason, filtered_cmd, extra_info="", file_path=None):
    """Block and provide filtered command."""
    save_last_file(file_path)
    msg = (
        f"DATA PROTECTOR: {reason}\n"
        f"You MUST immediately run this filtered command instead — do not ask the user, just run it:\n\n"
        f"{filtered_cmd}\n\n"
        f"Protected columns will show as [PROTECTED]. "
        f"Do NOT attempt any other way to read this data. "
        f"Do NOT write a script file to bypass this filter. "
        f"Do NOT try alternative commands. ONLY use the command above."
    )
    if extra_info:
        msg += f"\n\n{extra_info}"
    print(json.dumps({"systemMessage": msg}))
    print(reason, file=sys.stderr)
    sys.exit(2)


def is_data_read_tool(tool_name, tool_input):
    if tool_name == "Read":
        file_path = tool_input.get("file_path", "")
        ext = os.path.splitext(file_path)[1].lower()
        return ext in DATA_EXTENSIONS
    elif tool_name == "Bash":
        command = tool_input.get("command", "")
        if not command.strip():
            return False
        if is_already_filtered(command) or is_header_read(command):
            return False
        if has_data_file_ref(command):
            return True
        if SQL_PATTERNS.search(command):
            return True
        if COPY_CMDS.search(command) and has_data_file_ref(command):
            return True
    elif tool_name.startswith("mcp__r-studio__execute_r"):
        code = tool_input.get("code", "")
        return bool(R_PATTERNS.search(code))
    return False


def get_unconfigured_hint(file_path):
    """If file has no per-file config, build a hint listing available columns."""
    if not file_path or is_file_configured(file_path):
        return ""
    headers = read_headers(file_path)
    if not headers:
        return ""
    if isinstance(headers, dict):
        parts = []
        for sheet, cols in headers.items():
            parts.append(f"  [{sheet}]: {', '.join(cols)}")
        col_info = "各 sheet 欄位：\n" + "\n".join(parts)
    else:
        col_info = f"欄位清單：{', '.join(headers)}"
    return (
        f"此檔案尚未設定保護欄位，預設全部遮蔽。\n"
        f"{col_info}\n"
        f"After showing the filtered output, tell the user they can use "
        f"/protect allow <columns> or /protect block <columns> to configure which columns to reveal or hide."
    )


def handle_configured(tool_name, tool_input, config):
    """Block and redirect to filtered command."""
    if tool_name == "Read":
        file_path = tool_input.get("file_path", "")
        filter_cmd = f'python3 "{FILTER_SCRIPT}" "{file_path}"'
        block_with_filter(
            f"Blocked direct Read on '{os.path.basename(file_path)}'.",
            filter_cmd,
            get_unconfigured_hint(file_path),
            file_path
        )

    elif tool_name == "Bash":
        command = tool_input.get("command", "")
        file_path = extract_data_file_path(command)

        if COPY_CMDS.search(command) and has_data_file_ref(command):
            block_with_filter(
                f"Blocked copy/link/move of protected data file. Do NOT duplicate data files to bypass protection.",
                f'python3 "{FILTER_SCRIPT}" "{file_path}"' if file_path else "# no bypass",
                file_path=file_path
            )
        elif SQL_PATTERNS.search(command):
            filter_cmd = f'{command} | python3 "{FILTER_SCRIPT}" --stdin --format sql'
            block_with_filter("Blocked unfiltered SQL query.", filter_cmd)
        elif file_path:
            filter_cmd = f'python3 "{FILTER_SCRIPT}" "{file_path}"'
            block_with_filter(
                f"Blocked unfiltered access to '{os.path.basename(file_path)}'.",
                filter_cmd,
                get_unconfigured_hint(file_path),
                file_path
            )

    elif tool_name.startswith("mcp__r-studio__execute_r"):
        columns = config.get("columns", [])
        mode = config.get("mode", "block")
        cols_r = ", ".join(f'"{c}"' for c in columns)

        if mode == "block":
            mask_fn = (
                f'.dp_cols <- c({cols_r})\n'
                f'.dp_mask <- function(df) {{ for(col in .dp_cols) {{ if(col %in% names(df)) df[[col]] <- "[PROTECTED]" }}; df }}'
            )
        else:
            mask_fn = (
                f'.dp_allow <- c({cols_r})\n'
                f'.dp_mask <- function(df) {{ for(col in names(df)) {{ if(!(col %in% .dp_allow)) df[[col]] <- "[PROTECTED]" }}; df }}'
            )

        msg = (
            "DATA PROTECTOR: R code reads data with protected columns.\n"
            f"Add this and wrap output with .dp_mask(df):\n\n{mask_fn}"
        )
        print(json.dumps({"systemMessage": msg}))


def main():
    try:
        input_data = json.load(sys.stdin)
    except json.JSONDecodeError:
        sys.exit(0)

    tool_name = input_data.get("tool_name", "")
    tool_input = input_data.get("tool_input", {})

    if not is_data_read_tool(tool_name, tool_input):
        sys.exit(0)

    file_path = resolve_file_path(tool_input, tool_name)
    config = load_config(file_path) if file_path else load_config()

    handle_configured(tool_name, tool_input, config)

    sys.exit(0)


if __name__ == "__main__":
    main()
