#!/usr/bin/env python3
"""Core column filtering engine for data-protector plugin.

Usage:
  filter.py <file_path>          Read file, output with protected columns masked
  filter.py --stdin              Read stdin (piped output), filter columns
  filter.py --stdin --format sql Parse SQL-formatted output
"""
import csv
import io
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_manager import load_config, get_protected_indices, is_protected

MASK = "[PROTECTED]"

DATA_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".parquet"}


def detect_delimiter(line):
    for delim in ["\t", ",", "|", ";"]:
        if delim in line:
            return delim
    return ","


def filter_csv(text, config, delimiter=None):
    lines = text.strip().split("\n")
    if not lines:
        return text

    if delimiter is None:
        delimiter = detect_delimiter(lines[0])

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return text

    headers = rows[0]
    protected = get_protected_indices(headers, config)
    if not protected:
        return text

    out = io.StringIO()
    writer = csv.writer(out, delimiter=delimiter)
    for i, row in enumerate(rows):
        if i == 0:
            writer.writerow(row)
        else:
            masked = list(row)
            for idx in protected:
                if idx < len(masked):
                    masked[idx] = MASK
            writer.writerow(masked)
    return out.getvalue()


def filter_json_data(text, config):
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return text

    if isinstance(data, list) and data and isinstance(data[0], dict):
        for record in data:
            for key in list(record.keys()):
                if is_protected(key, config):
                    record[key] = MASK
        return json.dumps(data, indent=2, ensure_ascii=False)
    elif isinstance(data, dict):
        for key in list(data.keys()):
            if is_protected(key, config):
                data[key] = MASK
        return json.dumps(data, indent=2, ensure_ascii=False)
    return text


def filter_sql_output(text, config):
    lines = text.split("\n")
    if len(lines) < 2:
        return text

    header_line = None
    separator_line = None
    header_idx = -1

    for i, line in enumerate(lines):
        if re.match(r"^[\s|]*-+[\s|+-]*-+", line):
            separator_line = i
            if i > 0:
                header_line = i - 1
            break
        if "|" in line and not re.match(r"^\s*\+", line):
            potential_headers = [h.strip() for h in line.split("|") if h.strip()]
            if len(potential_headers) >= 2 and all(re.match(r"^[a-zA-Z_]\w*$", h) for h in potential_headers):
                header_line = i
                header_idx = i

    if header_line is None:
        return text

    header_text = lines[header_line]
    if "|" in header_text:
        headers = [h.strip() for h in header_text.split("|")]
        headers = [h for h in headers if h]
        delim = "|"
    else:
        headers = header_text.split()
        delim = None

    protected = get_protected_indices(headers, config)
    if not protected:
        return text

    result = []
    data_start = (separator_line + 1) if separator_line is not None else (header_line + 1)

    for i, line in enumerate(lines):
        if i <= header_line or (separator_line is not None and i == separator_line):
            result.append(line)
        elif i >= data_start and line.strip() and not re.match(r"^\s*\+", line) and not re.match(r"^\s*\(?\d+ rows?\)?", line):
            if delim == "|":
                parts = line.split("|")
                real_parts = []
                prefix = ""
                suffix = ""
                for j, p in enumerate(parts):
                    if j == 0 and not p.strip():
                        prefix = p
                        continue
                    if j == len(parts) - 1 and not p.strip():
                        suffix = p
                        continue
                    real_parts.append(p)

                for idx in protected:
                    if idx < len(real_parts):
                        orig = real_parts[idx]
                        width = len(orig)
                        real_parts[idx] = MASK.center(width) if width > len(MASK) else MASK

                result.append(prefix + "|".join(real_parts) + suffix)
            else:
                result.append(line)
        else:
            result.append(line)

    return "\n".join(result)


def filter_cat_numbered(text, config):
    """Handle `cat -n` style output (line numbers + tab + content)."""
    lines = text.split("\n")
    if not lines:
        return text

    numbered_pattern = re.compile(r"^(\s*\d+\t)(.*)$")
    first_content = None
    for line in lines:
        m = numbered_pattern.match(line)
        if m:
            first_content = m.group(2)
            break
        elif line.strip():
            first_content = line
            break

    if first_content is None:
        return text

    delimiter = detect_delimiter(first_content)
    is_numbered = bool(numbered_pattern.match(lines[0])) if lines[0].strip() else False

    if not is_numbered:
        return filter_csv(text, config, delimiter)

    content_lines = []
    prefixes = []
    for line in lines:
        m = numbered_pattern.match(line)
        if m:
            prefixes.append(m.group(1))
            content_lines.append(m.group(2))
        else:
            prefixes.append("")
            content_lines.append(line)

    filtered_content = filter_csv("\n".join(content_lines), config, delimiter)
    filtered_lines = filtered_content.split("\n")

    result = []
    for i, fline in enumerate(filtered_lines):
        prefix = prefixes[i] if i < len(prefixes) else ""
        result.append(prefix + fline)
    return "\n".join(result)


def filter_xlsx(file_path, config):
    try:
        import openpyxl
    except ImportError:
        import subprocess
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
            capture_output=True
        )
        import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        protected = get_protected_indices(headers, config)

        print(f"=== Sheet: {sheet_name} ({len(rows)-1} rows x {len(headers)} cols) ===")
        print("\t".join(headers))
        for row in rows[1:]:
            vals = []
            for i, v in enumerate(row):
                if i in protected:
                    vals.append(MASK)
                else:
                    vals.append(str(v) if v is not None else "")
            print("\t".join(vals))
        print()
    wb.close()


def filter_file(file_path, config):
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        filter_xlsx(file_path, config)
        return

    if ext == ".parquet":
        print(f"[DATA PROTECTOR] Parquet not supported directly. Use pandas to read.", file=sys.stderr)
        sys.exit(1)

    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    if ext == ".json":
        print(filter_json_data(content, config))
    elif ext in (".csv", ".tsv"):
        delimiter = "\t" if ext == ".tsv" else ","
        print(filter_csv(content, config, delimiter))
    else:
        print(filter_csv(content, config))


def filter_stdin(config, fmt=None):
    text = sys.stdin.read()
    if not text.strip():
        print(text, end="")
        return

    if fmt == "sql":
        print(filter_sql_output(text, config))
    elif fmt == "json":
        print(filter_json_data(text, config))
    else:
        try:
            json.loads(text)
            print(filter_json_data(text, config))
            return
        except (json.JSONDecodeError, ValueError):
            pass

        if re.search(r"^[\s|]*-+[\s|+-]*-+", text, re.MULTILINE):
            print(filter_sql_output(text, config))
        else:
            print(filter_cat_numbered(text, config))


def main():
    file_path = None
    if len(sys.argv) > 1 and sys.argv[1] != "--stdin":
        file_path = sys.argv[1]

    config = load_config(file_path)

    if "--stdin" in sys.argv:
        fmt = None
        if "--format" in sys.argv:
            fmt_idx = sys.argv.index("--format") + 1
            if fmt_idx < len(sys.argv):
                fmt = sys.argv[fmt_idx]
        filter_stdin(config, fmt)
    elif file_path:
        filter_file(file_path, config)
    else:
        print("Usage: filter.py <file_path> | filter.py --stdin [--format sql|json]", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
